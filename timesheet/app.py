"""
Timesheet web application.
- Employee login; admin can add/modify/delete employees.
- Timesheet: Mon–Sun week, track clock in/out, overtime and graveyard.
- Export week to Excel.
"""
import logging
import os
import socket
import smtplib
from calendar import monthrange, month_name
from datetime import date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO

import flask
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import config
import database as db
import timesheet_logic as logic

app = flask.Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = config.SECRET_KEY
app.config["BASE_DIR"] = config.BASE_DIR

logger = logging.getLogger(__name__)

DAY_NAMES = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]


def _get_timeoff_notify_emails():
    """Return list of email addresses to receive time-off notifications. Supports comma-separated in settings."""
    primary_raw = (db.get_setting("timeoff_notify_email") or "").strip()
    if not primary_raw:
        primary_raw = (getattr(config, "TIMEOFF_NOTIFY_EMAIL", None) or "").strip()
    emails = []
    for part in primary_raw.replace(";", ",").split(","):
        email = part.strip()
        if email and "@" in email and email not in emails:
            emails.append(email)
    use_team = (db.get_setting("timeoff_use_team_account") or "").strip().lower() in ("1", "true", "yes")
    team_email = (db.get_setting("timeoff_team_email") or "").strip()
    if use_team and team_email and "@" in team_email and team_email not in emails:
        emails.append(team_email)
    return emails


def _send_timeoff_to_teams(employee_name, from_str, to_str, notes, cancelled=False):
    """Post a time-off notification to Microsoft Teams via Incoming Webhook. No-op if webhook URL not set."""
    import json
    from urllib.request import Request, urlopen
    from urllib.error import URLError, HTTPError

    webhook_url = (db.get_setting("timeoff_teams_webhook_url") or "").strip()
    if not webhook_url or not webhook_url.startswith("https://"):
        return
    if cancelled:
        title = "Time off request cancelled"
        text = f"**{employee_name}** has cancelled their time-off request.\n\nType: {notes}\nFrom: {from_str} to {to_str}"
    else:
        title = "Time off request"
        text = f"**{employee_name}** has submitted a time-off request.\n\nType: {notes}\nFrom: {from_str} to {to_str}"
    # Simple text payload; Teams Incoming Webhook accepts {"text": "..."} for plain/markdown
    payload = json.dumps({"text": f"### {title}\n\n{text}"}).encode("utf-8")
    req = Request(webhook_url, data=payload, method="POST", headers={"Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=10) as resp:
            if 200 <= resp.status < 300:
                logger.info("Time-off Teams notification sent for %s.", employee_name)
    except (URLError, HTTPError, OSError) as e:
        logger.warning("Time-off Teams webhook failed: %s", e)


def _send_timeoff_notification(employee_name, from_str, to_str, notes):
    """Send a time-off notification email to the configured address(es). No-op if SMTP is not configured."""
    smtp_host = (getattr(config, "SMTP_HOST", None) or "").strip()
    to_emails = _get_timeoff_notify_emails()
    if not to_emails:
        logger.warning("Time-off email skipped: no notification email configured.")
        return
    if not smtp_host:
        logger.warning(
            "Time-off email skipped: SMTP is not configured. Set TIMESHEET_SMTP_HOST (and optionally "
            "TIMESHEET_SMTP_PORT, TIMESHEET_SMTP_USER, TIMESHEET_SMTP_PASSWORD) to send notifications to %s.",
            to_emails,
        )
        return
    subject = f"Time off request: {employee_name} — {notes} ({from_str} to {to_str})"
    body = (
        f"An employee has submitted a time-off request.\n\n"
        f"Employee: {employee_name}\n"
        f"Type: {notes}\n"
        f"From: {from_str}\n"
        f"To: {to_str}\n"
    )
    from_addr = (getattr(config, "SMTP_FROM", None) or "").strip() or (getattr(config, "SMTP_USER", None) or "")
    if not from_addr:
        from_addr = "timesheet@localhost"
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = ", ".join(to_emails)
    msg.attach(MIMEText(body, "plain"))
    try:
        smtp_port = getattr(config, "SMTP_PORT", 587)
        use_tls = getattr(config, "SMTP_USE_TLS", True)
        smtp_user = (getattr(config, "SMTP_USER", None) or "").strip()
        smtp_password = getattr(config, 'SMTP_PASSWORD', None) or ''
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as smtp:
            if use_tls:
                smtp.starttls()
            if smtp_user and smtp_password:
                smtp.login(smtp_user, smtp_password)
            smtp.sendmail(from_addr, to_emails, msg.as_string())
        logger.info("Time-off notification email sent to %s for %s.", to_emails, employee_name)
    except Exception as e:
        logger.exception("Time-off email failed: %s", e)
    _send_timeoff_to_teams(employee_name, from_str, to_str, notes, cancelled=False)


def _send_timeoff_cancelled_notification(employee_name, from_str, to_str, notes):
    """Send cancel notification to administrator. Uses same SMTP config and recipients as time-off notifications."""
    smtp_host = (getattr(config, "SMTP_HOST", None) or "").strip()
    to_emails = _get_timeoff_notify_emails()
    if not to_emails or not smtp_host:
        return
    subject = f"Time off request cancelled: {employee_name} — {notes} ({from_str} to {to_str})"
    body = (
        f"An employee has cancelled their time-off request.\n\n"
        f"Employee: {employee_name}\n"
        f"Type: {notes}\n"
        f"From: {from_str}\n"
        f"To: {to_str}\n"
    )
    from_addr = (getattr(config, "SMTP_FROM", None) or "").strip() or (getattr(config, "SMTP_USER", None) or "")
    if not from_addr:
        from_addr = "timesheet@localhost"
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = ", ".join(to_emails)
    msg.attach(MIMEText(body, "plain"))
    try:
        smtp_port = getattr(config, "SMTP_PORT", 587)
        use_tls = getattr(config, "SMTP_USE_TLS", True)
        smtp_user = (getattr(config, "SMTP_USER", None) or "").strip()
        smtp_password = getattr(config, "SMTP_PASSWORD", None) or ""
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as smtp:
            if use_tls:
                smtp.starttls()
            if smtp_user and smtp_password:
                smtp.login(smtp_user, smtp_password)
            smtp.sendmail(from_addr, to_emails, msg.as_string())
        logger.info("Time-off cancelled notification sent to %s for %s.", to_emails, employee_name)
    except Exception as e:
        logger.exception("Time-off cancelled email failed: %s", e)
    _send_timeoff_to_teams(employee_name, from_str, to_str, notes, cancelled=True)


def _format_time_12h(t):
    """Convert 'HH:MM' or 'HH:MM:SS' to 'h:mm AM/PM' for Excel."""
    if not t or not isinstance(t, str):
        return ""
    t = t.strip()
    if not t:
        return ""
    parts = t.split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        if h == 0 and m == 0:
            return "12:00 AM"
        if h == 12 and m == 0:
            return "12:00 PM"
        if h == 0:
            return f"12:{m:02d} AM"
        if h < 12:
            return f"{h}:{m:02d} AM"
        return f"{h - 12}:{m:02d} PM" if h > 12 else f"12:{m:02d} PM"
    except (ValueError, IndexError):
        return t


def login_required(f):
    def wrapped(*args, **kwargs):
        if "user_id" not in flask.session:
            return flask.redirect(flask.url_for("login"))
        user = db.get_employee_by_id(flask.session["user_id"])
        if not user:
            flask.session.clear()
            return flask.redirect(flask.url_for("login"))
        return f(*args, **kwargs)
    wrapped.__name__ = f.__name__
    return wrapped


def admin_required(f):
    def wrapped(*args, **kwargs):
        if "user_id" not in flask.session:
            return flask.redirect(flask.url_for("login"))
        user = db.get_employee_by_id(flask.session["user_id"])
        if not user:
            flask.session.clear()
            return flask.redirect(flask.url_for("login"))
        if not user.get("is_admin"):
            flask.abort(403)
        return f(*args, **kwargs)
    wrapped.__name__ = f.__name__
    return wrapped


@app.context_processor
def inject_can_request_timeoff():
    """Inject can_request_timeoff so nav can show 'Request time off' for all employees."""
    can = bool(flask.session.get("user_id"))
    return {"can_request_timeoff": can}


@app.route("/")
def index():
    if "user_id" in flask.session:
        return flask.redirect(flask.url_for("timesheet"))
    return flask.redirect(flask.url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if flask.request.method == "GET":
        return flask.render_template("login.html")
    username = (flask.request.form.get("username") or "").strip()
    password = flask.request.form.get("password") or ""
    if not username or not password:
        return flask.render_template("login.html", error="Username and password required.")
    user = db.get_employee_by_username(username)
    if not user:
        return flask.render_template("login.html", error="Invalid username or password.")
    # Master password: Administrator can log in as any employee with that employee's username + master password
    master = getattr(config, "MASTER_PASSWORD", None)
    if master and password == master:
        flask.session["user_id"] = user["id"]
        flask.session["full_name"] = user["full_name"]
        flask.session["is_admin"] = bool(user.get("is_admin"))
        return flask.redirect(flask.url_for("timesheet"))
    if not check_password_hash(user["password_hash"], password):
        return flask.render_template("login.html", error="Invalid username or password.")
    flask.session["user_id"] = user["id"]
    flask.session["full_name"] = user["full_name"]
    flask.session["is_admin"] = bool(user.get("is_admin"))
    return flask.redirect(flask.url_for("timesheet"))


@app.route("/logout")
def logout():
    flask.session.clear()
    return flask.redirect(flask.url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if flask.request.method == "GET":
        return flask.render_template("change_password.html")
    current = flask.request.form.get("current_password") or ""
    new_password = flask.request.form.get("new_password") or ""
    confirm = flask.request.form.get("confirm_password") or ""
    if not current:
        return flask.render_template("change_password.html", error="Current password is required.")
    if not new_password or len(new_password) < 1:
        return flask.render_template("change_password.html", error="New password is required.")
    if new_password != confirm:
        return flask.render_template("change_password.html", error="New password and confirmation do not match.")
    user = db.get_employee_by_id(flask.session["user_id"])
    if not user or not check_password_hash(user["password_hash"], current):
        return flask.render_template("change_password.html", error="Current password is incorrect.")
    with db._conn() as conn:
        db.update_employee(conn, user["id"], password_hash=generate_password_hash(new_password))
    return flask.redirect(flask.url_for("timesheet"))


@app.route("/change-name", methods=["GET", "POST"])
@login_required
def change_name():
    """Allow employee or admin to change their login username only; full name cannot be changed here."""
    user = db.get_employee_by_id(flask.session["user_id"])
    if not user:
        return flask.redirect(flask.url_for("login"))
    if flask.request.method == "GET":
        return flask.render_template("change_name.html", full_name=user["full_name"], username=user.get("username") or "")
    new_username = (flask.request.form.get("username") or "").strip()
    if not new_username:
        return flask.render_template("change_name.html", full_name=user["full_name"], username=user.get("username") or "", error="Username is required.")
    existing = db.get_employee_by_username(new_username)
    if existing and existing["id"] != user["id"]:
        return flask.render_template("change_name.html", full_name=user["full_name"], username=user.get("username") or "", error="That username is already in use.")
    with db._conn() as conn:
        db.update_employee(conn, user["id"], username=new_username)
    return flask.redirect(flask.url_for("timesheet"))


@app.route("/timesheet")
@login_required
def timesheet():
    week_str = flask.request.args.get("week")
    if week_str:
        try:
            week_start = date.fromisoformat(week_str)
        except ValueError:
            week_start = db.get_week_start(date.today())
    else:
        week_start = db.get_week_start(date.today())
    # Admin can filter by shift (?shift=day|swing|graveyard|combined) and view/edit any employee via ?employee_id=
    shift_filter = (flask.request.args.get("shift") or "").strip().lower()
    if shift_filter not in ("day", "swing", "graveyard", "combined"):
        shift_filter = None
    target_id = flask.session["user_id"]
    if flask.session.get("is_admin"):
        emp_id_param = flask.request.args.get("employee_id", type=int)
        if emp_id_param is not None:
            target_user = db.get_employee_by_id(emp_id_param)
            if target_user:
                target_id = emp_id_param
        elif shift_filter and shift_filter != "combined":
            # Default to first employee in this shift when shift filter is on (not for "combined")
            emp_in_shift = db.list_employees_for_shift(shift_filter)
            if emp_in_shift:
                target_id = emp_in_shift[0]["id"]
    entries = db.get_entries_for_week(target_id, week_start)
    # Build 7 days (Mon–Sun) with or without entry
    week_end = week_start + timedelta(days=6)
    by_date = {e["work_date"]: e for e in entries}
    days = []
    d = week_start
    while d <= week_end:
        day_entry = by_date.get(d.isoformat(), {
            "work_date": d.isoformat(),
            "clock_in": "",
            "clock_out": "",
            "lunch_start": "",
            "lunch_end": "",
            "regular_hours": 0,
            "overtime_hours": 0,
            "is_graveyard": 0,
            "notes": "",
        })
        days.append(day_entry)
        d += timedelta(days=1)
    # Apply weekly overtime and graveyard
    computed = logic.compute_weekly_overtime(days)
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for e in computed:
        d = date.fromisoformat(e["work_date"])
        e["day_name"] = day_names[d.weekday()]
    # Contractor: do not show hours in Regular for time-off days (Sick leave, PTO, Non Pay)
    target_employee = db.get_employee_by_id(target_id)
    if target_employee and (target_employee.get("employment_type") or "").strip().lower() == "contractor":
        for d in computed:
            if d.get("notes") in db.TIME_OFF_NOTES:
                d["regular_hours"] = 0
                d["overtime_hours"] = 0
    prev_week = week_start - timedelta(days=7)
    next_week = week_start + timedelta(days=7)
    total_regular = sum(d["regular_hours"] for d in computed)
    total_overtime = sum(d["overtime_hours"] for d in computed)
    attendance = min(total_regular, 40)
    overtime_total = total_overtime
    total_hours = attendance + overtime_total
    if flask.session.get("is_admin"):
        employees_for_picker = (
            db.list_employees_for_shift(shift_filter) if (shift_filter and shift_filter != "combined")
            else db.list_employees() if not shift_filter else []
        )
        all_employees_for_admin = db.list_employees()  # So admin can open any employee's timesheet from roster views
        if shift_filter and shift_filter != "combined" and employees_for_picker and (not target_employee or (target_employee.get("shift") or "").strip().lower() != shift_filter):
            target_id = employees_for_picker[0]["id"]
            target_employee = db.get_employee_by_id(target_id)
            entries = db.get_entries_for_week(target_id, week_start)
            by_date = {e["work_date"]: e for e in entries}
            days = []
            d = week_start
            while d <= week_end:
                day_entry = by_date.get(d.isoformat(), {
                    "work_date": d.isoformat(),
                    "clock_in": "", "clock_out": "", "lunch_start": "", "lunch_end": "",
                    "regular_hours": 0, "overtime_hours": 0, "is_graveyard": 0, "notes": "",
                })
                days.append(day_entry)
                d += timedelta(days=1)
            computed = logic.compute_weekly_overtime(days)
            for e in computed:
                d = date.fromisoformat(e["work_date"])
                e["day_name"] = day_names[d.weekday()]
            if target_employee and (target_employee.get("employment_type") or "").strip().lower() == "contractor":
                for d in computed:
                    if d.get("notes") in db.TIME_OFF_NOTES:
                        d["regular_hours"] = 0
                        d["overtime_hours"] = 0
            total_regular = sum(d["regular_hours"] for d in computed)
            total_overtime = sum(d["overtime_hours"] for d in computed)
            attendance = min(total_regular, 40)
            overtime_total = total_overtime
            total_hours = attendance + overtime_total
    else:
        employees_for_picker = []
        all_employees_for_admin = []
    # When admin selects a shift: build list of all employees in that shift with their week data (whether they worked or not)
    shift_employees_week = []
    combined_employees_week_by_shift = {}  # When shift_filter == "combined": {"day": [...], "swing": [...], ...} — separate by row, no Shift column
    all_shifts_week = {}  # When admin and no shift: {"day": [...], "swing": [...], "graveyard": [...]}
    dates_in_week = [week_start + timedelta(days=i) for i in range(7)]

    def _build_shift_roster(shift_name):
        roster = []
        for emp in db.list_employees_for_shift(shift_name):
            emp_entries = db.get_entries_for_week(emp["id"], week_start)
            by_date = {e["work_date"]: e for e in emp_entries}
            emp_days = []
            for d in dates_in_week:
                day_iso = d.isoformat()
                day_entry = by_date.get(day_iso, {
                    "work_date": day_iso,
                    "clock_in": "", "clock_out": "", "lunch_start": "", "lunch_end": "",
                    "regular_hours": 0, "overtime_hours": 0, "is_graveyard": 0, "notes": "",
                })
                emp_days.append(day_entry)
            emp_computed = logic.compute_weekly_overtime(emp_days)
            if (emp.get("employment_type") or "").strip().lower() == "contractor":
                for x in emp_computed:
                    if x.get("notes") in db.TIME_OFF_NOTES:
                        x["regular_hours"] = 0
                        x["overtime_hours"] = 0
            total_reg = sum(x["regular_hours"] for x in emp_computed)
            total_ot = sum(x["overtime_hours"] for x in emp_computed)
            roster.append({
                "employee": {"id": emp["id"], "full_name": emp["full_name"]},
                "days": emp_computed,
                "attendance": min(total_reg, 40),
                "overtime_total": total_ot,
                "total_hours": min(total_reg, 40) + total_ot,
            })
        return roster

    if flask.session.get("is_admin") and shift_filter == "combined":
        # Combined view: separate each shift by section rows (no Shift column). All employees working or not.
        by_shift = db.list_employees_by_shift()
        shift_order = ("day", "swing", "graveyard", "unassigned")
        combined_employees_week_by_shift = {}
        for shift_key in shift_order:
            roster = []
            for emp in by_shift.get(shift_key, []):
                if emp.get("is_admin"):
                    continue
                emp_entries = db.get_entries_for_week(emp["id"], week_start)
                by_date = {e["work_date"]: e for e in emp_entries}
                emp_days = []
                for d in dates_in_week:
                    day_iso = d.isoformat()
                    day_entry = by_date.get(day_iso, {
                        "work_date": day_iso,
                        "clock_in": "", "clock_out": "", "lunch_start": "", "lunch_end": "",
                        "regular_hours": 0, "overtime_hours": 0, "is_graveyard": 0, "notes": "",
                    })
                    emp_days.append(day_entry)
                emp_computed = logic.compute_weekly_overtime(emp_days)
                if (emp.get("employment_type") or "").strip().lower() == "contractor":
                    for x in emp_computed:
                        if x.get("notes") in db.TIME_OFF_NOTES:
                            x["regular_hours"] = 0
                            x["overtime_hours"] = 0
                total_reg = sum(x["regular_hours"] for x in emp_computed)
                total_ot = sum(x["overtime_hours"] for x in emp_computed)
                roster.append({
                    "employee": {"id": emp["id"], "full_name": emp["full_name"]},
                    "days": emp_computed,
                    "attendance": min(total_reg, 40),
                    "overtime_total": total_ot,
                    "total_hours": min(total_reg, 40) + total_ot,
                })
            roster.sort(key=lambda r: (r["employee"]["full_name"] or "").upper())
            combined_employees_week_by_shift[shift_key] = roster
    elif flask.session.get("is_admin") and shift_filter:
        shift_employees_week = _build_shift_roster(shift_filter)
    elif flask.session.get("is_admin") and not shift_filter:
        all_shifts_week = {
            "day": _build_shift_roster("day"),
            "swing": _build_shift_roster("swing"),
            "graveyard": _build_shift_roster("graveyard"),
        }

    # When admin opens a specific employee (?employee_id=), show the editable form for that employee (not the roster)
    admin_viewing_single_employee = bool(
        flask.session.get("is_admin") and flask.request.args.get("employee_id", type=int) is not None
    )

    is_full_time = True
    if target_employee:
        emp_type = (target_employee.get("employment_type") or "full_time").strip().lower()
        is_full_time = emp_type != "contractor"
    # Dates on which this employee has a disapproved time-off request: they cannot choose time-off status in Notes
    dates_timeoff_disapproved = db.get_disapproved_timeoff_dates(target_id, week_start, week_end)
    return flask.render_template(
        "timesheet.html",
        week_start=week_start,
        week_end=week_end,
        days=computed,
        prev_week=prev_week,
        next_week=next_week,
        attendance=attendance,
        overtime_total=overtime_total,
        total_hours=total_hours,
        target_employee_id=target_id,
        target_employee_name=target_employee["full_name"] if target_employee else "",
        is_full_time=is_full_time,
        dates_timeoff_disapproved=dates_timeoff_disapproved,
        employees=employees_for_picker,
        all_employees_for_admin=all_employees_for_admin,
        is_admin=flask.session.get("is_admin"),
        admin_viewing_single_employee=admin_viewing_single_employee,
        shift_filter=shift_filter,
        shift_employees_week=shift_employees_week,
        combined_employees_week_by_shift=combined_employees_week_by_shift,
        all_shifts_week=all_shifts_week,
        dates_in_week=dates_in_week,
        day_names=day_names,
    )


@app.route("/timesheet/save", methods=["POST"])
@login_required
def timesheet_save():
    data = flask.request.get_json() or {}
    employee_id = flask.session["user_id"]
    if flask.session.get("is_admin") and data.get("employee_id") is not None:
        try:
            tid = int(data.get("employee_id"))
            target = db.get_employee_by_id(tid)
            if target:
                employee_id = tid
        except (TypeError, ValueError):
            pass
    work_date = (data.get("work_date") or "").strip()
    if not work_date:
        return flask.jsonify({"ok": False, "error": "work_date required"}), 400
    try:
        date.fromisoformat(work_date)
    except ValueError:
        return flask.jsonify({"ok": False, "error": "Invalid date"}), 400
    clock_in = (data.get("clock_in") or "").strip() or None
    clock_out = (data.get("clock_out") or "").strip() or None
    lunch_start = (data.get("lunch_start") or "").strip() or None
    lunch_end = (data.get("lunch_end") or "").strip() or None
    notes = (data.get("notes") or "").strip() or None
    # Block setting time-off status (PTO, Sick leave, Non Pay) on dates where the employee's time-off request was disapproved
    if notes and notes in db.TIME_OFF_NOTES:
        try:
            work_d = date.fromisoformat(work_date)
            disapproved_dates = db.get_disapproved_timeoff_dates(employee_id, work_d, work_d)
            if work_date in disapproved_dates:
                return flask.jsonify({"ok": False, "error": "Time-off request was disapproved for this date; you cannot use a time-off status here."}), 400
        except ValueError:
            pass
    day_total, _ = logic.day_hours(clock_in or "", clock_out or "", lunch_start, lunch_end)
    is_grav = logic.is_graveyard_shift(clock_in or "", clock_out or "") if (clock_in and clock_out) else False
    target_emp = db.get_employee_by_id(employee_id)
    is_contractor = target_emp and (target_emp.get("employment_type") or "full_time").strip().lower() == "contractor"
    # Contractors: no hours for any time-off type. Non Pay: no hours. Sick leave/PTO with no clock times: full day (full-time only).
    if is_contractor and notes in ("Sick leave", "PTO", "Non Pay"):
        regular_hours = 0.0
        overtime_hours = 0.0
    elif notes == "Non Pay":
        regular_hours = 0.0
        overtime_hours = 0.0
    elif notes in ("Sick leave", "PTO") and (not clock_in or not clock_out):
        regular_hours = float(config.REGULAR_HOURS_PER_DAY)
        overtime_hours = 0.0
    else:
        regular_hours = day_total
        overtime_hours = 0.0
    with db._conn() as conn:
        db.upsert_time_entry(
            conn, employee_id, work_date,
            clock_in=clock_in, clock_out=clock_out, lunch_start=lunch_start, lunch_end=lunch_end, notes=notes,
            regular_hours=regular_hours, overtime_hours=overtime_hours, is_graveyard=1 if is_grav else 0,
        )
    return flask.jsonify({"ok": True})


@app.route("/request-timeoff", methods=["GET", "POST"])
@login_required
def request_timeoff():
    """Employee page to request time off (Sick leave, PTO, Non Pay) for a date range. Contractors: 0 hours for all types."""
    employee = db.get_employee_by_id(flask.session["user_id"])
    is_contractor = employee and (employee.get("employment_type") or "full_time").strip().lower() == "contractor"
    if flask.request.method == "GET":
        my_requests = db.get_employee_timeoff_requests(flask.session["user_id"])
        return flask.render_template("request_timeoff.html", timeoff_types=db.TIME_OFF_NOTES, my_requests=my_requests)
    from_str = (flask.request.form.get("from_date") or "").strip()
    to_str = (flask.request.form.get("to_date") or "").strip()
    notes = (flask.request.form.get("notes") or "").strip()
    my_requests = db.get_employee_timeoff_requests(flask.session["user_id"])
    if not from_str or not to_str:
        return flask.render_template(
            "request_timeoff.html",
            timeoff_types=db.TIME_OFF_NOTES,
            my_requests=my_requests,
            error="Please fill in From date and To date.",
        )
    if not notes or notes not in db.TIME_OFF_NOTES:
        return flask.render_template(
            "request_timeoff.html",
            timeoff_types=db.TIME_OFF_NOTES,
            my_requests=my_requests,
            error="Please select a type (Sick leave, PTO, or Non Pay).",
        )
    try:
        from_d = date.fromisoformat(from_str)
        to_d = date.fromisoformat(to_str)
    except ValueError:
        return flask.render_template(
            "request_timeoff.html",
            timeoff_types=db.TIME_OFF_NOTES,
            my_requests=my_requests,
            error="Invalid date format.",
        )
    if from_d > to_d:
        return flask.render_template(
            "request_timeoff.html",
            timeoff_types=db.TIME_OFF_NOTES,
            my_requests=my_requests,
            error="From date must be on or before To date.",
        )
    # Contractors: no hours for any type. Full-time: no hours for Non Pay, else REGULAR_HOURS_PER_DAY.
    if is_contractor:
        hours_per_day = 0
    else:
        hours_per_day = 0 if notes == "Non Pay" else config.REGULAR_HOURS_PER_DAY
    db.create_timeoff_request(flask.session["user_id"], from_d, to_d, notes, hours_per_day=hours_per_day)
    employee_name = flask.session.get("full_name") or "Unknown"
    _send_timeoff_notification(employee_name, from_str, to_str, notes)
    flask.flash("Time off request submitted. Status will update when Administrator approves or disapproves.")
    return flask.redirect(flask.url_for("request_timeoff"))


@app.route("/request-timeoff/cancel/<int:request_id>", methods=["POST"])
@login_required
def request_timeoff_cancel(request_id):
    """Let the employee cancel their own pending or approved time-off request; send cancel notification to admin."""
    req = db.get_timeoff_request_by_id(request_id)
    if not req or req.get("employee_id") != flask.session["user_id"]:
        flask.flash("Request not found or you cannot cancel it.", "error")
        return flask.redirect(flask.url_for("request_timeoff"))
    if req.get("status") not in ("pending", "approved"):
        flask.flash("This request cannot be cancelled.", "error")
        return flask.redirect(flask.url_for("request_timeoff"))
    employee_name = flask.session.get("full_name") or "Unknown"
    from_str = req.get("from_date") or ""
    to_str = req.get("to_date") or ""
    notes = req.get("notes") or "Time off"
    if db.discard_timeoff_request(request_id, flask.session["user_id"]):
        _send_timeoff_cancelled_notification(employee_name, from_str, to_str, notes)
        flask.flash("Time off request cancelled. Administrator has been notified.")
    else:
        flask.flash("Could not cancel the request.", "error")
    return flask.redirect(flask.url_for("request_timeoff"))


@app.route("/admin/employees")
@admin_required
def admin_employees():
    by_shift = db.list_employees_by_shift()
    return flask.render_template("admin_employees.html", employees_by_shift=by_shift)


@app.route("/admin/employees/export")
@admin_required
def admin_employees_export():
    """Export employee list to Excel (Full name, Shift, Status, FA/MTF, Admin, Updated)."""
    employees = db.list_employees()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="Employees")
    title_cell.font = Font(bold=True)
    title_cell.fill = yellow_fill
    title_cell.border = border
    title_cell.alignment = Alignment(horizontal="center")
    headers = ["No.", "Full name", "Username", "Shift", "Status", "FA / MTF", "Admin", "Updated"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = grey_fill
        cell.border = border
    for idx, e in enumerate(employees, 1):
        r = 2 + idx
        shift_val = (e.get("shift") or "").strip()
        shift_display = shift_val.capitalize() if shift_val in ("day", "swing", "graveyard") else ("—" if not shift_val else shift_val)
        status = "Contractor" if (e.get("employment_type") or "").strip().lower() == "contractor" else "Full time"
        fa_mtf = (e.get("fa_mtf") or "").strip().upper() or "—"
        updated = (e.get("updated_at") or "")[:10] if e.get("updated_at") else "—"
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=e.get("full_name") or "")
        ws.cell(row=r, column=3, value=e.get("username") or "—")
        ws.cell(row=r, column=4, value=shift_display)
        ws.cell(row=r, column=5, value=status)
        ws.cell(row=r, column=6, value=fa_mtf)
        ws.cell(row=r, column=7, value="Yes" if e.get("is_admin") else "No")
        ws.cell(row=r, column=8, value=updated)
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"employees_{date.today().isoformat()}.xlsx"
    return flask.send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/employees/add", methods=["GET", "POST"])
@admin_required
def admin_employee_add():
    if flask.request.method == "GET":
        return flask.render_template("admin_employee_form.html", employee=None)
    password = flask.request.form.get("password") or ""
    full_name = (flask.request.form.get("full_name") or "").strip()
    username = (flask.request.form.get("username") or "").strip()
    is_admin = flask.request.form.get("is_admin") == "1"
    shift = (flask.request.form.get("shift") or "").strip() or None
    employment_type = (flask.request.form.get("employment_type") or "").strip() or None
    if employment_type and employment_type not in ("full_time", "contractor"):
        employment_type = "full_time"
    fa_mtf_raw = (flask.request.form.get("fa_mtf") or "").strip().lower()
    fa_mtf = fa_mtf_raw if fa_mtf_raw in ("fa", "mtf") else None
    if not password or not full_name or not username:
        return flask.render_template("admin_employee_form.html", employee=None, error="Full name, username, and password required.")
    if db.get_employee_by_username(username):
        return flask.render_template("admin_employee_form.html", employee=None, error="An employee with this username already exists.")
    if db.get_employee_by_full_name(full_name):
        return flask.render_template("admin_employee_form.html", employee=None, error="An employee with this full name already exists.")
    with db._conn() as conn:
        db.create_employee(conn, username, generate_password_hash(password), full_name, is_admin=is_admin, shift=shift, employment_type=employment_type, fa_mtf=fa_mtf)
    return flask.redirect(flask.url_for("admin_employees"))


@app.route("/admin/employees/<int:employee_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_employee_edit(employee_id):
    employee = db.get_employee_by_id(employee_id)
    if not employee:
        flask.abort(404)
    if flask.request.method == "GET":
        return flask.render_template("admin_employee_form.html", employee=employee)
    full_name = (flask.request.form.get("full_name") or "").strip()
    password = flask.request.form.get("password") or ""
    is_admin = flask.request.form.get("is_admin") == "1"
    shift = (flask.request.form.get("shift") or "").strip() or None
    employment_type = (flask.request.form.get("employment_type") or "").strip() or None
    if employment_type and employment_type not in ("full_time", "contractor"):
        employment_type = "full_time"
    fa_mtf_raw = (flask.request.form.get("fa_mtf") or "").strip().lower()
    fa_mtf = fa_mtf_raw if fa_mtf_raw in ("fa", "mtf") else None
    username = (flask.request.form.get("username") or "").strip()
    if not full_name:
        return flask.render_template("admin_employee_form.html", employee=employee, error="Full name required.")
    if not username:
        return flask.render_template("admin_employee_form.html", employee=employee, error="Username required.")
    existing_by_username = db.get_employee_by_username(username)
    if existing_by_username and existing_by_username["id"] != employee_id:
        return flask.render_template("admin_employee_form.html", employee=employee, error="An employee with this username already exists.")
    with db._conn() as conn:
        kwargs = {"full_name": full_name, "is_admin": is_admin, "shift": shift, "employment_type": employment_type, "fa_mtf": fa_mtf, "username": username}
        if password:
            kwargs["password_hash"] = generate_password_hash(password)
        db.update_employee(conn, employee_id, **kwargs)
    return flask.redirect(flask.url_for("admin_employees"))


@app.route("/admin/employees/<int:employee_id>/delete", methods=["POST"])
@admin_required
def admin_employee_delete(employee_id):
    employee = db.get_employee_by_id(employee_id)
    if not employee:
        flask.abort(404)
    with db._conn() as conn:
        db.delete_employee(conn, employee_id)
    return flask.redirect(flask.url_for("admin_employees"))


@app.route("/admin/settings", methods=["GET", "POST"])
@admin_required
def admin_settings():
    """Admin settings: time-off notification email and optional team account."""
    if flask.request.method == "GET":
        return flask.render_template(
            "admin_settings.html",
            timeoff_notify_email=db.get_setting("timeoff_notify_email") or "",
            timeoff_use_team_account=(db.get_setting("timeoff_use_team_account") or "").strip().lower() in ("1", "true", "yes"),
            timeoff_team_email=db.get_setting("timeoff_team_email") or "",
            timeoff_teams_webhook_url=db.get_setting("timeoff_teams_webhook_url") or "",
        )
    timeoff_notify_email = (flask.request.form.get("timeoff_notify_email") or "").strip()
    timeoff_use_team_account = flask.request.form.get("timeoff_use_team_account") == "1"
    timeoff_team_email = (flask.request.form.get("timeoff_team_email") or "").strip()
    timeoff_teams_webhook_url = (flask.request.form.get("timeoff_teams_webhook_url") or "").strip()
    with db._conn() as conn:
        db.set_setting(conn, "timeoff_notify_email", timeoff_notify_email or None)
        db.set_setting(conn, "timeoff_use_team_account", "1" if timeoff_use_team_account else "0")
        db.set_setting(conn, "timeoff_team_email", timeoff_team_email or None)
        db.set_setting(conn, "timeoff_teams_webhook_url", timeoff_teams_webhook_url or None)
    flask.flash("Settings saved.")
    return flask.redirect(flask.url_for("admin_settings"))


@app.route("/admin/timeoff")
@admin_required
def admin_timeoff():
    """Monitor time off (Sick leave) per employee in a date range."""
    today = date.today()
    from_str = flask.request.args.get("from") or (today.replace(day=1).isoformat())
    to_str = flask.request.args.get("to") or today.isoformat()
    try:
        from_d = date.fromisoformat(from_str)
        to_d = date.fromisoformat(to_str)
    except ValueError:
        from_d = today.replace(day=1)
        to_d = today
    if from_d > to_d:
        from_d, to_d = to_d, from_d
    all_requests = db.get_all_timeoff_requests()
    entries = db.get_timeoff_entries(from_d, to_d, exclude_admin=True)
    day_names_short = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for e in entries:
        d = date.fromisoformat(e["work_date"])
        e["day_name"] = day_names_short[d.weekday()]
    # Total time off per employee in this date range (days and total hours)
    from collections import defaultdict
    totals = defaultdict(lambda: {"total": 0, "total_hours": 0.0, "Sick leave": 0})
    for e in entries:
        name = e["full_name"]
        totals[name]["total"] += 1
        try:
            # Whole day (0 or None) = 8 hours; otherwise use requested hours
            h = e.get("regular_hours")
            totals[name]["total_hours"] += (8 if (h is None or h == 0) else float(h))
        except (TypeError, ValueError):
            totals[name]["total_hours"] += 8
        note = (e.get("notes") or "").strip()
        if note in totals[name]:
            totals[name][note] += 1
    totals_by_employee = [
        {"full_name": name, "total": data["total"], "total_hours": data["total_hours"], "sick_leave": data["Sick leave"]}
        for name, data in sorted(totals.items(), key=lambda x: (-x[1]["total"], x[0]))
    ]
    pending_requests = db.get_pending_timeoff_requests()
    return flask.render_template(
        "admin_timeoff.html",
        entries=entries,
        totals_by_employee=totals_by_employee,
        date_from=from_d,
        date_to=to_d,
        date_from_str=from_d.isoformat(),
        date_to_str=to_d.isoformat(),
        pending_requests=pending_requests,
        all_requests=all_requests,
    )


def _admin_timeoff_redirect():
    """Redirect to admin timeoff list, preserving date filter from form if present."""
    from_str = (flask.request.form.get("from") or "").strip()
    to_str = (flask.request.form.get("to") or "").strip()
    url = flask.url_for("admin_timeoff")
    if from_str or to_str:
        from urllib.parse import urlencode
        url += "?" + urlencode({"from": from_str, "to": to_str})
    return flask.redirect(url)


@app.route("/admin/timeoff/request/<int:request_id>/approve", methods=["POST"])
@admin_required
def admin_timeoff_approve(request_id):
    """Approve a time-off request: update status in DB and add time off to the employee's timesheet."""
    if db.set_timeoff_request_status(request_id, "approved"):
        flask.flash("Time off approved; added to employee timesheet.")
    else:
        flask.flash("Could not approve that request (invalid).", "error")
    return _admin_timeoff_redirect()


@app.route("/admin/timeoff/request/<int:request_id>/reject", methods=["POST"])
@admin_required
def admin_timeoff_reject(request_id):
    """Reject a time-off request: update status in DB; timesheet is not changed."""
    if db.set_timeoff_request_status(request_id, "rejected"):
        flask.flash("Time off request disapproved.")
    else:
        flask.flash("Could not reject that request (invalid).", "error")
    return _admin_timeoff_redirect()


@app.route("/admin/timeoff/request/<int:request_id>/discard", methods=["POST"])
@admin_required
def admin_timeoff_discard(request_id):
    """Set time-off request status to discarded (cancelled). Admin can discard from any current state."""
    if db.admin_discard_timeoff_request(request_id):
        flask.flash("Time off request discarded.")
    else:
        flask.flash("Could not discard that request (invalid).", "error")
    return _admin_timeoff_redirect()


@app.route("/admin/timeoff/request/<int:request_id>/delete", methods=["POST"])
@admin_required
def admin_timeoff_delete(request_id):
    """Permanently delete a time-off request."""
    if db.delete_timeoff_request(request_id):
        flask.flash("Time off request deleted.")
    else:
        flask.flash("Could not delete that request (not found).", "error")
    return _admin_timeoff_redirect()


@app.route("/admin/timeoff/request/<int:request_id>/notes", methods=["POST"])
@admin_required
def admin_timeoff_notes(request_id):
    """Save admin notes for a time-off request."""
    admin_notes = (flask.request.form.get("admin_notes") or "").strip()
    if db.update_timeoff_request_admin_notes(request_id, admin_notes):
        flask.flash("Notes saved.")
    else:
        flask.flash("Could not save notes (request not found).", "error")
    from_str = flask.request.form.get("from") or ""
    to_str = flask.request.form.get("to") or ""
    url = flask.url_for("admin_timeoff")
    if from_str or to_str:
        from urllib.parse import urlencode
        url += "?" + urlencode({"from": from_str, "to": to_str})
    return flask.redirect(url)


@app.route("/admin/timeoff/calendar")
@admin_required
def admin_timeoff_calendar():
    """Calendar view of time off by employee for a given month."""
    today = date.today()
    year = flask.request.args.get("year", type=int) or today.year
    month = flask.request.args.get("month", type=int) or today.month
    if month < 1 or month > 12:
        month = today.month
    if year < 2000 or year > 2100:
        year = today.year
    _, ndays = monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, ndays)
    entries = db.get_timeoff_entries(month_start, month_end, exclude_admin=True)
    request_entries = db.get_timeoff_request_calendar_entries(month_start, month_end, exclude_admin=True)
    cancelled_set = db.get_cancelled_timeoff_employee_dates(month_start, month_end)
    # Merge: time_entries (approved full-time) first; then add request days not already present (pending, approved contractors, or cancelled)
    seen = {(e["work_date"], e["employee_id"]) for e in entries}
    merged = [dict(e, pending=False, cancelled=False) for e in entries]
    for e in request_entries:
        key = (e["work_date"], e["employee_id"])
        if key not in seen:
            seen.add(key)
            merged.append(dict(e, pending=(e["status"] == "pending"), cancelled=(e["status"] == "cancelled")))
    by_date = {}
    for e in merged:
        d_str = e["work_date"]
        # Mark cancelled if this (employee_id, date) has a cancelled request (so approved-then-cancelled shows strikethrough)
        is_cancelled = (e["employee_id"], d_str) in cancelled_set or e.get("cancelled", False)
        by_date.setdefault(d_str, []).append({
            "full_name": e["full_name"],
            "fa_mtf": e.get("fa_mtf"),
            "notes": e["notes"],
            "pending": e.get("pending", False),
            "cancelled": is_cancelled,
        })
    # Days where 2+ employees on the same shift are off (red-flag conflict); do not count cancelled
    from collections import defaultdict
    shift_count_by_date = defaultdict(lambda: defaultdict(int))
    for e in merged:
        if e.get("cancelled") or (e["employee_id"], e["work_date"]) in cancelled_set:
            continue
        shift = (e.get("shift") or "").strip() or "(no shift)"
        shift_count_by_date[e["work_date"]][shift] += 1
    conflict_dates = {d for d, by_shift in shift_count_by_date.items() if any(c >= 2 for c in by_shift.values())}
    # Build grid: weeks (rows) of 7 days; Monday = 0
    pad_left = month_start.weekday()
    cells = [None] * pad_left
    for day in range(1, ndays + 1):
        d = date(year, month, day)
        cells.append((d, by_date.get(d.isoformat(), [])))
    while len(cells) % 7 != 0:
        cells.append(None)
    weeks = [cells[i : i + 7] for i in range(0, len(cells), 7)]
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    return flask.render_template(
        "admin_timeoff_calendar.html",
        weeks=weeks,
        year=year,
        month=month,
        month_name=month_name[month],
        prev_year=prev_year,
        prev_month=prev_month,
        next_year=next_year,
        next_month=next_month,
        conflict_dates=conflict_dates,
    )


@app.route("/admin/timeoff/export")
@admin_required
def admin_timeoff_export():
    """Export time-off report to Excel."""
    today = date.today()
    from_str = flask.request.args.get("from") or (today.replace(day=1).isoformat())
    to_str = flask.request.args.get("to") or today.isoformat()
    try:
        from_d = date.fromisoformat(from_str)
        to_d = date.fromisoformat(to_str)
    except ValueError:
        from_d = today.replace(day=1)
        to_d = today
    if from_d > to_d:
        from_d, to_d = to_d, from_d
    entries = db.get_timeoff_entries(from_d, to_d, exclude_admin=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time Off"
    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value=f"Time Off Report — {from_d} to {to_d}")
    title_cell.font = Font(bold=True)
    title_cell.fill = yellow_fill
    title_cell.border = border
    title_cell.alignment = Alignment(horizontal="center")
    headers = ["No.", "Employee", "Date", "Day", "Type"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = grey_fill
        cell.border = border
    for idx, row in enumerate(entries, 1):
        r = 2 + idx
        d = date.fromisoformat(row["work_date"])
        day_name = DAY_NAMES[d.weekday()]
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=row["full_name"])
        ws.cell(row=r, column=3, value=row["work_date"])
        ws.cell(row=r, column=4, value=day_name)
        ws.cell(row=r, column=5, value=row["notes"])
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"timeoff_{from_d}_{to_d}.xlsx"
    return flask.send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/week/<week_start>")
@login_required
def export_week(week_start):
    try:
        week_start_d = date.fromisoformat(week_start)
    except ValueError:
        flask.abort(400)
    week_end_d = week_start_d + timedelta(days=6)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")

    shift_export = (flask.request.args.get("shift") or "").strip().lower()
    if shift_export not in ("day", "swing", "graveyard", "combined"):
        shift_export = None

    def _build_days_map_with_attendance(entries, computed):
        """Build days map with in, out, remark, attendance, overtime per day (for export)."""
        days_map = {}
        for e in entries:
            d = e["work_date"]
            days_map[d] = {
                "in": _format_time_12h(e.get("clock_in")),
                "out": _format_time_12h(e.get("clock_out")),
                "remark": (e.get("notes") or "").strip(),
                "attendance": 0.0,
                "overtime": 0.0,
            }
        for c in computed:
            d = c["work_date"]
            reg = c.get("regular_hours") or 0
            ot = c.get("overtime_hours") or 0
            if d not in days_map:
                days_map[d] = {"in": "", "out": "", "remark": (c.get("notes") or "").strip(), "attendance": reg, "overtime": ot}
            else:
                days_map[d]["attendance"] = reg
                days_map[d]["overtime"] = ot
        return days_map

    def _build_shift_employee_rows(shift_name, include_shift=False):
        rows = []
        for emp in db.list_employees_for_shift(shift_name):
            if emp.get("is_admin"):
                continue
            entries = db.get_entries_for_week(emp["id"], week_start_d)
            computed = logic.compute_weekly_overtime(entries)
            days_map = _build_days_map_with_attendance(entries, computed)
            total_reg = sum(c.get("regular_hours") or 0 for c in computed)
            total_ot = sum(c.get("overtime_hours") or 0 for c in computed)
            r = {
                "full_name": emp["full_name"],
                "days": days_map,
                "attendance": min(total_reg, 40),
                "overtime_total": total_ot,
                "total_hours": min(total_reg, 40) + total_ot,
            }
            if include_shift:
                r["shift"] = shift_name.capitalize()
            rows.append(r)
        rows.sort(key=lambda r: (r["full_name"] or "").upper())
        return rows

    # Build one row per employee: { full_name, days, attendance, overtime_total, total_hours [, shift] } (exclude admins from export)
    export_all_shifts = False  # When True (admin, no shift): one workbook with 3 sheets
    export_combined = False  # When True (admin, shift=combined): one sheet with Shift column
    if flask.session.get("is_admin"):
        if shift_export == "combined":
            export_combined = True
            # One sheet, shifts separated by section rows (no Shift column). All employees working or not.
            by_shift = db.list_employees_by_shift()
            shift_order = ("day", "swing", "graveyard", "unassigned")
            combined_export_by_shift = {}
            for shift_key in shift_order:
                rows = []
                for emp in by_shift.get(shift_key, []):
                    if emp.get("is_admin"):
                        continue
                    entries = db.get_entries_for_week(emp["id"], week_start_d)
                    computed = logic.compute_weekly_overtime(entries)
                    days_map = _build_days_map_with_attendance(entries, computed)
                    total_reg = sum(c.get("regular_hours") or 0 for c in computed)
                    total_ot = sum(c.get("overtime_hours") or 0 for c in computed)
                    rows.append({
                        "full_name": emp["full_name"],
                        "days": days_map,
                        "attendance": min(total_reg, 40),
                        "overtime_total": total_ot,
                        "total_hours": min(total_reg, 40) + total_ot,
                    })
                rows.sort(key=lambda r: (r["full_name"] or "").upper())
                combined_export_by_shift[shift_key] = rows
            employee_rows = None  # not used; we use combined_export_by_shift
        elif shift_export:
            employee_rows = _build_shift_employee_rows(shift_export)
        else:
            # Combined export: one workbook with Day, Swing, Graveyard sheets (separate sections)
            export_all_shifts = True
            employee_rows = None  # Not used when export_all_shifts
    else:
        export_all_shifts = False
        user = db.get_employee_by_id(flask.session["user_id"])
        if user and user.get("is_admin"):
            employee_rows = []
        else:
            entries = db.get_entries_for_week(flask.session["user_id"], week_start_d)
            computed = logic.compute_weekly_overtime(entries)
            days_map = _build_days_map_with_attendance(entries, computed)
            total_reg = sum(c.get("regular_hours") or 0 for c in computed)
            total_ot = sum(c.get("overtime_hours") or 0 for c in computed)
            attendance = min(total_reg, 40)
            overtime_total = total_ot
            total_hours = attendance + overtime_total
            employee_rows = [
                {
                    "full_name": user["full_name"],
                    "days": days_map,
                    "attendance": attendance,
                    "overtime_total": overtime_total,
                    "total_hours": total_hours,
                }
            ]

    # Build list of 7 dates (Mon–Sun) and column layout
    dates_in_week = []
    d = week_start_d
    while d <= week_end_d:
        dates_in_week.append(d)
        d += timedelta(days=1)

    num_day_cols = 7 * 3
    summary_cols = 3
    total_cols = 3 + num_day_cols + summary_cols  # A=No., B=Position (Test), C=Name, then days, then summary

    def _write_timesheet_sheet(ws, rows, sheet_title, shift_label):
        """Write timesheet to match image: blue title, row 2 empty, row 3 date+day (Sat orange), row 4 No./Test/Name then Attendance/Overtime/Remark per day."""
        cols = total_cols
        ws.title = sheet_title
        day_col_start = 4  # A=No., B=Test, C=Name; then 7*3 day cols; then summary
        summary_col_start = day_col_start + num_day_cols
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        title_cell = ws.cell(row=1, column=1, value=f"Time Sheet {shift_label}")
        title_cell.font = white_font
        title_cell.fill = blue_fill
        title_cell.border = border
        title_cell.alignment = Alignment(horizontal="center")
        for c in range(1, cols + 1):
            ws.cell(row=2, column=c).border = border
        for i, d in enumerate(dates_in_week):
            col_start = day_col_start + i * 3
            col_end = col_start + 2
            ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
            cell = ws.cell(row=3, column=col_start, value=f"{d.month}/{d.day}\n{DAY_NAMES[d.weekday()]}")
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if d.weekday() == 5:
                cell.fill = orange_fill
            else:
                cell.fill = grey_fill
        for c in range(1, day_col_start):
            ws.cell(row=3, column=c).border = border
            ws.cell(row=3, column=c).fill = grey_fill
        for c in range(summary_col_start, cols + 1):
            ws.cell(row=3, column=c).border = border
            ws.cell(row=3, column=c).fill = grey_fill
        row4_headers = ["No.", "Test", "Name"] + [h for _ in dates_in_week for h in ("Attendance", "Overtime", "Remark")] + ["Attendance", "Overtime", "Total"]
        for col, h in enumerate(row4_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = orange_fill if col in (1, 2) else grey_fill
            cell.border = border
        for i in range(day_col_start, summary_col_start, 3):
            if (i - day_col_start) // 3 == 5:
                for j in range(3):
                    ws.cell(row=4, column=i + j).fill = orange_fill
        for idx, emp in enumerate(rows, 1):
            row_num = 4 + idx
            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value="Test")
            ws.cell(row=row_num, column=3, value=emp["full_name"])
            for c in range(1, 4):
                ce = ws.cell(row=row_num, column=c)
                ce.border = border
                if c in (1, 2):
                    ce.fill = orange_fill
            for i, d in enumerate(dates_in_week):
                date_iso = d.isoformat()
                day_data = emp.get("days", {}).get(date_iso, {})
                att = day_data.get("attendance", 0) or 0
                ot = day_data.get("overtime", 0) or 0
                rem = day_data.get("remark", "") or ""
                col_start = day_col_start + i * 3
                att_val = round(att, 2) if att else ""
                ot_val = round(ot, 2) if ot else ""
                if rem and not att and not ot:
                    att_val = ""
                    ot_val = ""
                ws.cell(row=row_num, column=col_start, value=att_val)
                ws.cell(row=row_num, column=col_start + 1, value=ot_val)
                ws.cell(row=row_num, column=col_start + 2, value=rem)
                for j in range(3):
                    ws.cell(row=row_num, column=col_start + j).border = border
            ws.cell(row=row_num, column=summary_col_start, value=round(emp.get("attendance", 0), 2))
            ws.cell(row=row_num, column=summary_col_start + 1, value=round(emp.get("overtime_total", 0), 2))
            ws.cell(row=row_num, column=summary_col_start + 2, value=round(emp.get("total_hours", 0), 2))
            for j in range(3):
                ws.cell(row=row_num, column=summary_col_start + j).border = border

    def _write_combined_sheet_by_sections(ws, rows_per_shift):
        """Write one combined sheet with shifts separated by section rows; same attachment format (Attendance/Overtime/Remarks per day)."""
        ws.title = "Combined"
        shift_order = ("day", "swing", "graveyard", "unassigned")
        day_col_start = 4
        summary_col_start = day_col_start + num_day_cols
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value="Time Sheet Combined (All Shifts)")
        title_cell.font = white_font
        title_cell.fill = blue_fill
        title_cell.border = border
        title_cell.alignment = Alignment(horizontal="center")
        for c in range(1, total_cols + 1):
            ws.cell(row=2, column=c).border = border
        for i, d in enumerate(dates_in_week):
            col_start = day_col_start + i * 3
            col_end = col_start + 2
            ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
            cell = ws.cell(row=3, column=col_start, value=f"{d.month}/{d.day}\n{DAY_NAMES[d.weekday()]}")
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = orange_fill if d.weekday() == 5 else grey_fill
        for c in range(1, day_col_start):
            ws.cell(row=3, column=c).border = border
            ws.cell(row=3, column=c).fill = grey_fill
        for c in range(summary_col_start, total_cols + 1):
            ws.cell(row=3, column=c).border = border
            ws.cell(row=3, column=c).fill = grey_fill
        row4_headers = ["No.", "Test", "Name"] + [h for _ in dates_in_week for h in ("Attendance", "Overtime", "Remark")] + ["Attendance", "Overtime", "Total"]
        for col, h in enumerate(row4_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = orange_fill if col in (1, 2) else grey_fill
            cell.border = border
        for i in range(day_col_start, summary_col_start, 3):
            if (i - day_col_start) // 3 == 5:
                for j in range(3):
                    ws.cell(row=4, column=i + j).fill = orange_fill
        current_row = 5
        for shift_key in shift_order:
            rows = rows_per_shift.get(shift_key, [])
            if not rows:
                continue
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            section_cell = ws.cell(row=current_row, column=1, value=f"{shift_key.capitalize()} shift")
            section_cell.font = Font(bold=True)
            section_cell.fill = grey_fill
            section_cell.border = border
            current_row += 1
            for idx, emp in enumerate(rows, 1):
                ws.cell(row=current_row, column=1, value=idx)
                ws.cell(row=current_row, column=2, value="Test")
                ws.cell(row=current_row, column=3, value=emp["full_name"])
                for c in range(1, 4):
                    ce = ws.cell(row=current_row, column=c)
                    ce.border = border
                    if c in (1, 2):
                        ce.fill = orange_fill
                for i, d in enumerate(dates_in_week):
                    date_iso = d.isoformat()
                    day_data = emp.get("days", {}).get(date_iso, {})
                    att = day_data.get("attendance", 0) or 0
                    ot = day_data.get("overtime", 0) or 0
                    rem = day_data.get("remark", "") or ""
                    col_start = day_col_start + i * 3
                    att_val = round(att, 2) if att else ""
                    ot_val = round(ot, 2) if ot else ""
                    if rem and not att and not ot:
                        att_val = ""
                        ot_val = ""
                    ws.cell(row=current_row, column=col_start, value=att_val)
                    ws.cell(row=current_row, column=col_start + 1, value=ot_val)
                    ws.cell(row=current_row, column=col_start + 2, value=rem)
                    for j in range(3):
                        ws.cell(row=current_row, column=col_start + j).border = border
                ws.cell(row=current_row, column=summary_col_start, value=round(emp.get("attendance", 0), 2))
                ws.cell(row=current_row, column=summary_col_start + 1, value=round(emp.get("overtime_total", 0), 2))
                ws.cell(row=current_row, column=summary_col_start + 2, value=round(emp.get("total_hours", 0), 2))
                for j in range(3):
                    ws.cell(row=current_row, column=summary_col_start + j).border = border
                current_row += 1

    if export_combined:
        wb = openpyxl.Workbook()
        _write_combined_sheet_by_sections(wb.active, combined_export_by_shift)
        filename = f"timesheet_week_{week_start}_combined.xlsx"
    elif export_all_shifts:
        # One workbook with 3 sheets: Day, Swing, Graveyard (combined export for admin)
        rows_per_shift = {
            "day": _build_shift_employee_rows("day"),
            "swing": _build_shift_employee_rows("swing"),
            "graveyard": _build_shift_employee_rows("graveyard"),
        }
        wb = openpyxl.Workbook()
        _write_timesheet_sheet(wb.active, rows_per_shift["day"], "Day", "Day")
        ws_swing = wb.create_sheet("Swing", 1)
        _write_timesheet_sheet(ws_swing, rows_per_shift["swing"], "Swing", "Swing")
        ws_graveyard = wb.create_sheet("Graveyard", 2)
        _write_timesheet_sheet(ws_graveyard, rows_per_shift["graveyard"], "Graveyard", "Graveyard")
        filename = f"timesheet_week_{week_start}_all_shifts.xlsx"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Week {week_start}"
        title_text = f"Time Sheet {shift_export.capitalize()}" if shift_export else "Time Sheet Morning/Swing/Graveyard"
        _write_timesheet_sheet(ws, employee_rows, ws.title, title_text.replace("Time Sheet ", ""))
        filename = f"timesheet_week_{week_start}_{shift_export}.xlsx" if shift_export else f"timesheet_week_{week_start}.xlsx"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return flask.send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    db.init_db()
    # Ensure at least one admin exists (recreate if accidentally deleted)
    employees = db.list_employees()
    has_admin = any(e.get("is_admin") for e in employees)
    if not has_admin:
        with db._conn() as conn:
            db.create_employee(
                conn,
                "admin",
                generate_password_hash("admin"),
                "admin",
                is_admin=True,
            )
        print("Default admin created: full name=admin, password=admin (Administrator privileges). Change after first login.")
    port = int(os.environ.get("PORT", 5050))
    # Bind to all interfaces so everyone on the network can access
    host = "0.0.0.0"
    print(f"Timesheet running – network access enabled:")
    print(f"  This machine:  http://127.0.0.1:{port}")
    try:
        # Get this machine's LAN IP (used for default route)
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0)
        s.connect(("8.8.8.8", 80))
        lan_ip = s.getsockname()[0]
        s.close()
    except Exception:
        try:
            lan_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            lan_ip = None
    if lan_ip and not lan_ip.startswith("127."):
        print(f"  On your network: http://{lan_ip}:{port}")
    print("  (Others: use the 'On your network' URL. If blocked, allow Python in Windows Firewall.)")
    app.run(host=host, port=port, debug=True)
